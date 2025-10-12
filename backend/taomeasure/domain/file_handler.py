"""
File handling helpers for coordinate batch processing.
"""

from __future__ import annotations

import base64
import csv
import os
from io import BytesIO, StringIO
from typing import Any, Dict, Iterable, List, Optional, Tuple


class FileHandler:
    """Utilities for reading, validating and exporting coordinate files."""

    def __init__(self) -> None:
        self.supported_formats = [".txt", ".csv", ".dat", ".xls", ".xlsx"]

    def prepare_lines(
        self,
        file_content: Optional[str],
        *,
        filename: Optional[str] = None,
        encoding: str = "text",
    ) -> List[str]:
        """Decode uploaded payload into a list of normalised text lines."""

        if not file_content:
            return []

        if encoding.lower() == "base64":
            if not filename:
                raise ValueError("Base64 uploads require the original filename to infer the format.")
            extension = os.path.splitext(filename)[1].lower()
            binary = base64.b64decode(file_content)
            if extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                return self._read_xlsx_lines(binary)
            if extension in {".xls", ".xlt"}:
                return self._read_xls_lines(binary)
            text = binary.decode("utf-8", errors="ignore")
            return self._split_text_lines(text)

        return self._split_text_lines(file_content)

    def _split_text_lines(self, text: str) -> List[str]:
        """Split plain text into lines while normalising line endings."""

        normalized = text.replace("\r\n", "\n").replace("\r", "\n")
        return normalized.split("\n")

    def _read_xlsx_lines(self, data: bytes) -> List[str]:
        """Load the first worksheet of an XLSX file and produce tab separated lines."""

        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pragma: no cover - optional dependency
            raise RuntimeError("读取 XLSX 文件需要预先安装 openpyxl 库。") from exc

        workbook = load_workbook(BytesIO(data), data_only=True)
        sheet = workbook.active
        rows = (tuple(cell for cell in row) for row in sheet.iter_rows(values_only=True))
        return self._rows_to_lines(rows)

    def _read_xls_lines(self, data: bytes) -> List[str]:
        """Load the first worksheet of a legacy XLS file and produce tab separated lines."""

        try:
            import xlrd  # type: ignore
        except ImportError as exc:  # pragma: no cover - optional dependency
            raise RuntimeError("读取 XLS 文件需要预先安装 xlrd==1.2.0 库。") from exc

        workbook = xlrd.open_workbook(file_contents=data)
        sheet = workbook.sheet_by_index(0)
        rows = (sheet.row_values(idx) for idx in range(sheet.nrows))
        return self._rows_to_lines(rows)

    def _rows_to_lines(self, rows: Iterable[Iterable[Any]]) -> List[str]:
        """Convert a sequence of worksheet rows to tab separated strings."""

        lines: List[str] = []
        for row in rows:
            values = [self._stringify_cell(value) for value in row]
            while values and values[-1] == "":
                values.pop()
            if not any(values):
                continue
            lines.append("\t".join(values))
        return lines

    @staticmethod
    def _stringify_cell(value: Any) -> str:
        """Format worksheet cell values into readable strings."""

        if value is None:
            return ""
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return f"{value:.12g}"
        return str(value).strip()

    @staticmethod
    def parse_coordinate_value(value_str: str) -> Optional[float]:
        """Parse a coordinate token supporting decimal degrees or d.mmss notation."""

        try:
            value_str = value_str.strip()
            if not value_str:
                return None
            if "." in value_str:
                parts = value_str.split(".")
                if len(parts) == 2 and len(parts[1]) == 4:
                    degrees = float(parts[0])
                    mmss = parts[1]
                    minutes = float(mmss[:2])
                    seconds = float(mmss[2:])
                    decimal = abs(degrees) + minutes / 60.0 + seconds / 3600.0
                    if degrees < 0:
                        decimal = -decimal
                    return decimal
            return float(value_str)
        except (ValueError, IndexError):
            return None

    def parse_coordinate_file(self, file_lines: List[str], file_type: str) -> Dict[str, Any]:
        """Parse lines into structured point dictionaries based on file type."""

        lines = file_lines or []
        points: List[Dict[str, Any]] = []

        if file_type == "geodetic_to_cartesian":
            for line in lines:
                tokens = line.strip().split()
                if len(tokens) >= 4:
                    lat = self.parse_coordinate_value(tokens[1])
                    lon = self.parse_coordinate_value(tokens[2])
                    height = self.parse_coordinate_value(tokens[3])
                    if None not in (lat, lon, height):
                        points.append(
                            {
                                "name": tokens[0],
                                "lat": lat,
                                "lon": lon,
                                "height": height,
                            }
                        )

        elif file_type == "cartesian_to_geodetic":
            for line in lines:
                tokens = line.strip().split()
                if len(tokens) >= 4:
                    x = self.parse_coordinate_value(tokens[1])
                    y = self.parse_coordinate_value(tokens[2])
                    z = self.parse_coordinate_value(tokens[3])
                    if None not in (x, y, z):
                        points.append(
                            {
                                "name": tokens[0],
                                "x": x,
                                "y": y,
                                "z": z,
                            }
                        )

        elif file_type == "gauss_forward":
            for line in lines:
                tokens = line.strip().split()
                if len(tokens) >= 3:
                    lat = self.parse_coordinate_value(tokens[1])
                    lon = self.parse_coordinate_value(tokens[2])
                    if None not in (lat, lon):
                        points.append({"name": tokens[0], "lat": lat, "lon": lon})

        elif file_type == "gauss_inverse":
            for line in lines:
                tokens = line.strip().split()
                if len(tokens) >= 3:
                    x = self.parse_coordinate_value(tokens[1])
                    y = self.parse_coordinate_value(tokens[2])
                    if None not in (x, y):
                        points.append({"name": tokens[0], "x": x, "y": y})

        elif file_type == "four_param_control":
            for line in lines:
                tokens = line.strip().split("\t")
                if len(tokens) >= 5:
                    src_x = self.parse_coordinate_value(tokens[1])
                    src_y = self.parse_coordinate_value(tokens[2])
                    dst_x = self.parse_coordinate_value(tokens[3])
                    dst_y = self.parse_coordinate_value(tokens[4])
                    if None not in (src_x, src_y, dst_x, dst_y):
                        points.append(
                            {
                                "name": tokens[0],
                                "src_x": src_x,
                                "src_y": src_y,
                                "dst_x": dst_x,
                                "dst_y": dst_y,
                            }
                        )

        elif file_type == "four_param_unknown":
            for line in lines:
                tokens = line.strip().split("\t")
                if len(tokens) >= 3:
                    x = self.parse_coordinate_value(tokens[1])
                    y = self.parse_coordinate_value(tokens[2])
                    if None not in (x, y):
                        points.append({"name": tokens[0], "x": x, "y": y})

        elif file_type == "seven_param_control":
            for line in lines:
                tokens = line.strip().split("\t")
                if len(tokens) >= 7:
                    coords = [self.parse_coordinate_value(tokens[i]) for i in range(1, 7)]
                    if all(value is not None for value in coords):
                        points.append(
                            {
                                "name": tokens[0],
                                "src_x": coords[0],
                                "src_y": coords[1],
                                "src_z": coords[2],
                                "dst_x": coords[3],
                                "dst_y": coords[4],
                                "dst_z": coords[5],
                            }
                        )

        elif file_type == "seven_param_unknown":
            for line in lines:
                tokens = line.strip().split("\t")
                if len(tokens) >= 4:
                    values = [self.parse_coordinate_value(tokens[i]) for i in range(1, 4)]
                    if all(value is not None for value in values):
                        points.append(
                            {
                                "name": tokens[0],
                                "x": values[0],
                                "y": values[1],
                                "z": values[2],
                            }
                        )

        elif file_type == "zone_transform":
            for line in lines:
                if not line.strip():
                    continue
                if "," in line:
                    tokens = [token.strip() for token in line.split(",")]
                else:
                    tokens = line.strip().split()
                if len(tokens) >= 3:
                    x = self.parse_coordinate_value(tokens[1])
                    y = self.parse_coordinate_value(tokens[2])
                    if None not in (x, y):
                        point: Dict[str, Any] = {"name": tokens[0], "x": x, "y": y}
                        if len(tokens) >= 4:
                            h = self.parse_coordinate_value(tokens[3])
                            if h is not None:
                                point["height"] = h
                        points.append(point)

        return {"points": points, "count": len(points), "format": file_type}

    def export_results_to_text(self, results: List[Dict[str, Any]], result_type: str) -> str:
        """Format conversion results as a human readable text block."""

        output_lines: List[str] = []

        if result_type == "geodetic_to_cartesian":
            output_lines.append("点名\t\tX(m)\t\tY(m)\t\tZ(m)")
            output_lines.append("-" * 60)
            for item in results:
                if "error" in item:
                    output_lines.append(f"{item['name']}\t\t{item['error']}")
                    continue
                output_lines.append(
                    f"{item['name']}\t\t{item['x']:.3f}\t\t{item['y']:.3f}\t\t{item['z']:.3f}"
                )

        elif result_type == "cartesian_to_geodetic":
            output_lines.append("点名\t\t纬度(°)\t\t经度(°)\t\t大地高(m)")
            output_lines.append("-" * 70)
            for item in results:
                if "error" in item:
                    output_lines.append(f"{item['name']}\t\t{item['error']}")
                    continue
                output_lines.append(
                    f"{item['name']}\t\t{item['lat']:.8f}\t\t{item['lon']:.8f}\t\t{item['height']:.3f}"
                )

        elif result_type == "gauss_forward":
            output_lines.append("点名\t\tX(m)\t\tY(m)")
            output_lines.append("-" * 50)
            for item in results:
                if "error" in item:
                    output_lines.append(f"{item['name']}\t\t{item['error']}")
                    continue
                output_lines.append(f"{item['name']}\t\t{item['x']:.3f}\t\t{item['y']:.3f}")

        elif result_type == "gauss_inverse":
            output_lines.append("点名\t\t纬度(°)\t\t经度(°)")
            output_lines.append("-" * 60)
            for item in results:
                if "error" in item:
                    output_lines.append(f"{item['name']}\t\t{item['error']}")
                    continue
                output_lines.append(f"{item['name']}\t\t{item['lat']:.8f}\t\t{item['lon']:.8f}")

        elif result_type == "zone_transform":
            output_lines.append("点名\t\t原X(m)\t\t原Y(m)\t\t新X(m)\t\t新Y(m)\t\t备注")
            output_lines.append("-" * 80)
            for item in results:
                output_lines.append(
                    "\t\t".join(
                        [
                            str(item.get("name", "")),
                            f"{item.get('src_x', 0.0):.3f}",
                            f"{item.get('src_y', 0.0):.3f}",
                            f"{item.get('dst_x', 0.0):.3f}",
                            f"{item.get('dst_y', 0.0):.3f}",
                            item.get("new_zone", ""),
                        ]
                    )
                )

        return "\n".join(output_lines)

    def export_results_to_csv(self, results: List[Dict[str, Any]], result_type: str) -> str:
        """Format conversion results as CSV."""

        output = StringIO()
        writer = csv.writer(output)

        if result_type == "geodetic_to_cartesian":
            writer.writerow(["点名", "X(m)", "Y(m)", "Z(m)"])
            for item in results:
                writer.writerow(
                    [
                        item.get("name", ""),
                        f"{item.get('x', 0.0):.3f}",
                        f"{item.get('y', 0.0):.3f}",
                        f"{item.get('z', 0.0):.3f}",
                    ]
                )

        elif result_type == "cartesian_to_geodetic":
            writer.writerow(["点名", "纬度(°)", "经度(°)", "大地高(m)"])
            for item in results:
                writer.writerow(
                    [
                        item.get("name", ""),
                        f"{item.get('lat', 0.0):.8f}",
                        f"{item.get('lon', 0.0):.8f}",
                        f"{item.get('height', 0.0):.3f}",
                    ]
                )

        elif result_type == "gauss_forward":
            writer.writerow(["点名", "X(m)", "Y(m)"])
            for item in results:
                writer.writerow([
                    item.get("name", ""),
                    f"{item.get('x', 0.0):.3f}",
                    f"{item.get('y', 0.0):.3f}",
                ])

        elif result_type == "gauss_inverse":
            writer.writerow(["点名", "纬度(°)", "经度(°)"])
            for item in results:
                writer.writerow([
                    item.get("name", ""),
                    f"{item.get('lat', 0.0):.8f}",
                    f"{item.get('lon', 0.0):.8f}",
                ])

        elif result_type == "zone_transform":
            writer.writerow(["点名", "原X(m)", "原Y(m)", "新X(m)", "新Y(m)", "新带号"])
            for item in results:
                writer.writerow([
                    item.get("name", ""),
                    f"{item.get('src_x', 0.0):.3f}",
                    f"{item.get('src_y', 0.0):.3f}",
                    f"{item.get('dst_x', 0.0):.3f}",
                    f"{item.get('dst_y', 0.0):.3f}",
                    item.get("new_zone", ""),
                ])

        return output.getvalue()

    def validate_file_format(self, file_lines: List[str], expected_format: str) -> Tuple[bool, str]:
        """Basic structural validation for supported import formats."""

        lines = [line for line in (file_lines or []) if line and line.strip()]
        if not lines:
            return False, "文件内容为空"

        if expected_format == "geodetic_to_cartesian":
            for index, line in enumerate(lines, start=1):
                tokens = line.split()
                if len(tokens) < 4:
                    return False, f"第 {index} 行缺少经纬高数据"
                if any(self.parse_coordinate_value(tokens[i]) is None for i in (1, 2, 3)):
                    return False, f"第 {index} 行经纬高解析失败"

        elif expected_format == "cartesian_to_geodetic":
            for index, line in enumerate(lines, start=1):
                tokens = line.split()
                if len(tokens) < 4:
                    return False, f"第 {index} 行缺少 XYZ 数据"
                if any(self.parse_coordinate_value(tokens[i]) is None for i in (1, 2, 3)):
                    return False, f"第 {index} 行 XYZ 解析失败"

        elif expected_format in {"gauss_forward", "gauss_inverse"}:
            for index, line in enumerate(lines, start=1):
                tokens = line.split()
                if len(tokens) < 3:
                    return False, f"第 {index} 行字段数量不足"
                if any(self.parse_coordinate_value(tokens[i]) is None for i in (1, 2)):
                    return False, f"第 {index} 行坐标解析失败"

        elif expected_format in {
            "four_param_control",
            "four_param_unknown",
            "seven_param_control",
            "seven_param_unknown",
        }:
            for index, line in enumerate(lines, start=1):
                tokens = line.split("\t")
                min_count = 5 if "control" in expected_format else 3
                if "seven_param" in expected_format:
                    min_count = 7 if "control" in expected_format else 4
                if len(tokens) < min_count:
                    return False, f"第 {index} 行字段数量不足"
                for token in tokens[1:]:
                    if self.parse_coordinate_value(token) is None:
                        return False, f"第 {index} 行存在非法数字"

        return True, ""

    def get_file_info(self, file_lines: List[str]) -> Dict[str, Any]:
        """Analyse basic statistics of uploaded content."""

        lines = file_lines or []
        valid_lines = [line for line in lines if line.strip()]

        detected_format = "unknown"
        if valid_lines:
            first = valid_lines[0].strip()
            parts_space = first.split()
            parts_tab = first.split("\t")
            parts_comma = first.split(",")

            if len(parts_tab) > len(parts_space) and len(parts_tab) >= 5:
                detected_format = "parameter_transform"
            elif len(parts_comma) > 1:
                detected_format = "zone_transform"
            elif len(parts_space) == 4:
                try:
                    b = self.parse_coordinate_value(parts_space[1])
                    l = self.parse_coordinate_value(parts_space[2])
                    h = self.parse_coordinate_value(parts_space[3])
                    if None not in (b, l, h) and 0 <= (b or 0) <= 90 and 0 <= (l or 0) <= 360:
                        detected_format = "geodetic_to_cartesian"
                    else:
                        detected_format = "cartesian_to_geodetic"
                except Exception:  # pragma: no cover - best effort detection
                    pass
            elif len(parts_space) == 3:
                detected_format = "gauss_projection"

        return {
            "total_lines": len(lines),
            "valid_lines": len(valid_lines),
            "empty_lines": len(lines) - len(valid_lines),
            "detected_format": detected_format,
            "sample_line": valid_lines[0] if valid_lines else "",
        }
